import csv
import json
import re
from dataclasses import fields
from itertools import groupby
from operator import attrgetter
from typing import Iterable, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import (
    Color,
    Font,
    PatternFill,
)
from openpyxl.worksheet import Worksheet

from .models import Amendement


FIELDS = [field.name for field in fields(Amendement)]

FIELDS_NAMES = {
    'article': 'Nº article',
    'alinea': 'Alinéa',
    'num': 'Nº amdt ou sous-amdt',
    'auteur': 'Auteur(s)',
    'date_depot': 'Date de dépôt',
    'discussion_commune': 'Discussion commune ?',
    'identique': 'Identique ?',
}


DARK_BLUE = Color(rgb='00182848')
WHITE = Color(rgb='00FFFFFF')
GROUPS_COLORS = {
    'Les Républicains': '#2011e8',  # https://fr.wikipedia.org/wiki/Les_R%C3%A9publicains # noqa
    'UC': '#1e90ff',  # https://fr.wikipedia.org/wiki/Groupe_Union_centriste_(S%C3%A9nat) # noqa
    'SOCR': '#ff8080',  # https://fr.wikipedia.org/wiki/Groupe_socialiste_et_r%C3%A9publicain_(S%C3%A9nat) # noqa
    'RDSE': '#a38ebc',  # https://fr.wikipedia.org/wiki/Groupe_du_Rassemblement_d%C3%A9mocratique_et_social_europ%C3%A9en # noqa
    'Les Indépendants': '#30bfe9',  # https://fr.wikipedia.org/wiki/Groupe_Les_Ind%C3%A9pendants_%E2%80%93_R%C3%A9publique_et_territoires # noqa
    'NI': '#ffffff',  # https://fr.wikipedia.org/wiki/Non-inscrit
}


def write_csv(amendements: Iterable[Amendement], filename: str) -> int:
    nb_rows = 0
    with open(filename, 'w', encoding='utf-8') as file_:
        writer = csv.DictWriter(
            file_,
            fieldnames=FIELDS,
            delimiter=';',
            quoting=csv.QUOTE_MINIMAL,
        )
        writer.writeheader()
        for amendement in amendements:
            writer.writerow(amendement.asdict())
            nb_rows += 1
    return nb_rows


def write_xlsx(amendements: Iterable[Amendement], filename: str) -> int:
    wb = Workbook()
    ws = wb.active
    ws.title = "Amendements"

    _write_header_row(ws)
    nb_rows = _write_data_rows(ws, amendements)
    wb.save(filename)
    return nb_rows


def _write_header_row(ws: Worksheet) -> None:
    header_row = [
        FIELDS_NAMES.get(field, field.capitalize())
        for field in FIELDS
    ]
    for column, value in enumerate(header_row, 1):
        cell = ws.cell(row=1, column=column)
        cell.value = value
        cell.fill = PatternFill(
            patternType='solid',
            fgColor=DARK_BLUE,
        )
        cell.font = Font(
            color=WHITE,
            sz=8,
        )


def _write_data_rows(ws: Worksheet,
                     amendements: Iterable[Amendement]) -> int:
    nb_rows = 0
    for amend in amendements:
        values = tuple(amend.asdict().values())
        for column, value in enumerate(values, 1):
            cell = ws.cell(row=nb_rows + 2, column=column)
            cell.value = value
            cell.font = Font(sz=8)
        nb_rows += 1
    return nb_rows


def write_json_for_viewer(id_projet: int,
                          title: str,
                          amendements: Iterable[Amendement],
                          filename: str) -> int:
    data = _format_amendements(id_projet, title, amendements)
    with open(filename, 'w') as file_:
        json.dump(data, file_)
    return len(list(amendements))


def _format_amendements(id_projet: int,
                        title: str,
                        amendements: Iterable[Amendement]) -> list:
    sorted_amendements = sorted(amendements, key=attrgetter('article'))
    return [{
        "idProjet": id_projet,
        "libelle": title,
        "list": [
            _format_article(article, items)
            for article, items in groupby(
                sorted_amendements,
                key=attrgetter('article'),
            )
            if article
        ],
    }]


def _format_article(article: str,
                    amendements: Iterable[Amendement]) -> dict:
    id_, mult = _parse_article(article)
    return {
        "idArticle": id_,
        "etat": "",
        "multiplicatif": mult,
        "titre": "TODO",
        "feuilletJaune": f"jaune-{id_}{mult}.pdf",
        "document": f"article-{id_}{mult}.pdf",
        "amendements": [
            _format_amendement(amendement)
            for amendement in amendements
        ]
    }


ARTICLE_RE = re.compile(
    r'''^
        .*                      # some stuff before
        Article\s(?P<id>\d+)    # article number
        (?:\s(?P<mult>\w+))?    # bis, ter, etc.
        (?:\s.*)?               # some stuff after
        $
    ''', re.VERBOSE
)


def _parse_article(article: str) -> Tuple[Optional[int], str]:
    if article == '':
        return (None, '')
    mo = ARTICLE_RE.match(article)
    if mo is None:
        raise ValueError(f"Could not parse article number {article!r}")
    return int(mo.group('id')), mo.group('mult') or ""


def _format_amendement(amendement: Amendement) -> dict:
    return {
        "idAmendement": amendement.num_int,
        "etat": amendement.sort or '',
        "gouvernemental": amendement.gouvernemental,
        "auteurs": [
            {
                "auteur": amendement.auteur
            }
        ],
        "groupesParlementaires": [
            {
                "libelle": amendement.groupe or '',
                "couleur": GROUPS_COLORS.get(amendement.groupe, "#ffffff")
            }
        ],
        "document": f"{amendement.num_int:06}-00.pdf",
    }
